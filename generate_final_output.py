"""
Traceability Mapper v3 — Final output generator
Takes LLM verdicts + original data and produces the final traceability Excel.
"""

import json
import openpyxl
import pandas as pd
from datetime import datetime

OUTPUT_FILE = "traceability_output_v3.xlsx"

# Load evaluation data
with open('evaluation_pairs.json') as f:
    eval_pairs = json.load(f)

with open('all_verdicts.json') as f:
    all_verdicts = json.load(f)

# Load full 3P data for complete details
IGNORE_SHEETS = ["Sheet1", "Report Overview", "Summary"]
THREE_P_FILE = "vizio_3p_test_suites.xlsx"

wb = openpyxl.load_workbook(THREE_P_FILE, read_only=True)
three_p_full = {}  # id -> full data
for sheet_name in wb.sheetnames:
    if sheet_name in IGNORE_SHEETS:
        continue
    ws = wb[sheet_name]
    header_row = None
    header_map = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        row_vals = [str(v).strip().lower() if v else "" for v in row]
        if "id" in row_vals and "title" in row_vals:
            header_row = i
            for j, val in enumerate(row_vals):
                if val:
                    header_map[val] = j
            break
    if header_row is None:
        continue
    id_col = header_map.get("id")
    title_col = header_map.get("title")
    category_col = header_map.get("category")
    precond_col = header_map.get("preconds")
    steps_col = header_map.get("steps")
    steps_sep_col = header_map.get("steps separated")
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_list = list(row)
        tc_id = row_list[id_col] if id_col is not None and id_col < len(row_list) else None
        if not tc_id:
            continue
        tc_id_str = str(tc_id).strip()
        three_p_full[tc_id_str] = {
            "sheet": sheet_name,
            "id": tc_id_str,
            "title": str(row_list[title_col]).strip() if title_col is not None and title_col < len(row_list) and row_list[title_col] else "",
            "category": str(row_list[category_col]).strip() if category_col is not None and category_col < len(row_list) and row_list[category_col] else "",
            "preconditions": str(row_list[precond_col]).strip() if precond_col is not None and precond_col < len(row_list) and row_list[precond_col] else "",
            "steps": str(row_list[steps_col]).strip() if steps_col is not None and steps_col < len(row_list) and row_list[steps_col] else "",
            "steps_separated": str(row_list[steps_sep_col]).strip() if steps_sep_col is not None and steps_sep_col < len(row_list) and row_list[steps_sep_col] else "",
        }
wb.close()

# Build traceability rows
rows = []
for portal_id, portal_data in eval_pairs.items():
    portal_info = portal_data["portal"]
    verdicts = all_verdicts.get(portal_id, [])
    candidates = portal_data["candidates"]

    # Match verdicts to candidates by 3p_id
    matched_any = False
    for verdict in verdicts:
        if verdict["verdict"] == "NO":
            continue  # Skip NO verdicts

        three_p_id = str(verdict["3p_id"]).strip()
        # Get full 3P details
        three_p_info = three_p_full.get(three_p_id, {})
        if not three_p_info:
            # Try finding in candidates
            for cand in candidates:
                if str(cand["id"]).strip() == three_p_id:
                    three_p_info = cand
                    break

        # Map verdict to confidence
        if verdict["verdict"] == "YES":
            confidence = "HIGH_CONFIDENCE"
        else:  # PARTIAL
            conf = verdict.get("confidence", "MEDIUM_CONFIDENCE")
            if conf == "HIGH_CONFIDENCE":
                confidence = "MEDIUM_CONFIDENCE"  # PARTIAL can't be HIGH
            else:
                confidence = conf

        rows.append({
            "PORTAL Test Case ID": portal_info["id"],
            "PORTAL Category": portal_info["category"],
            "PORTAL Title": portal_info["title"],
            "PORTAL Description": portal_info["description"],
            "3P Sheet/Suite": three_p_info.get("sheet", ""),
            "3P Category": three_p_info.get("category", ""),
            "3P Test Case ID": three_p_id,
            "3P Title": three_p_info.get("title", ""),
            "3P Preconditions": three_p_info.get("preconditions", ""),
            "3P Steps": three_p_info.get("steps", ""),
            "3P Steps Separated": three_p_info.get("steps_separated", ""),
            "Confidence": confidence,
            "LLM Verdict": verdict["verdict"],
            "Reasoning": verdict.get("reason", ""),
        })
        matched_any = True

    if not matched_any:
        # NO_MATCH row
        rows.append({
            "PORTAL Test Case ID": portal_info["id"],
            "PORTAL Category": portal_info["category"],
            "PORTAL Title": portal_info["title"],
            "PORTAL Description": portal_info["description"],
            "3P Sheet/Suite": "",
            "3P Category": "",
            "3P Test Case ID": "",
            "3P Title": "",
            "3P Preconditions": "",
            "3P Steps": "",
            "3P Steps Separated": "",
            "Confidence": "NO_MATCH",
            "LLM Verdict": "NO_MATCH",
            "Reasoning": "No functionally equivalent 3P test case found in top 10 embedding candidates.",
        })

print(f"Total traceability rows: {len(rows)}")

# Generate executive summary stats
portal_ids = list(eval_pairs.keys())
total_portal = len(portal_ids)

coverage_stats = {}
for portal_id in portal_ids:
    verdicts = all_verdicts.get(portal_id, [])
    has_yes = any(v["verdict"] == "YES" for v in verdicts)
    has_partial = any(v["verdict"] == "PARTIAL" for v in verdicts)
    yes_count = sum(1 for v in verdicts if v["verdict"] == "YES")
    partial_count = sum(1 for v in verdicts if v["verdict"] == "PARTIAL")

    if has_yes:
        coverage_stats[portal_id] = "FULL"
    elif has_partial:
        coverage_stats[portal_id] = "PARTIAL"
    else:
        coverage_stats[portal_id] = "NONE"

full_coverage = sum(1 for v in coverage_stats.values() if v == "FULL")
partial_coverage = sum(1 for v in coverage_stats.values() if v == "PARTIAL")
no_coverage = sum(1 for v in coverage_stats.values() if v == "NONE")

# Unique 3P IDs mapped
mapped_3p_ids = set()
for r in rows:
    if r["Confidence"] != "NO_MATCH" and r["3P Test Case ID"]:
        mapped_3p_ids.add(r["3P Test Case ID"])

# Category stats
category_stats = {}
for portal_id in portal_ids:
    cat = eval_pairs[portal_id]["portal"]["category"]
    if cat not in category_stats:
        category_stats[cat] = {"total": 0, "full": 0, "partial": 0, "none": 0}
    category_stats[cat]["total"] += 1
    cov = coverage_stats[portal_id]
    if cov == "FULL":
        category_stats[cat]["full"] += 1
    elif cov == "PARTIAL":
        category_stats[cat]["partial"] += 1
    else:
        category_stats[cat]["none"] += 1

print(f"Full coverage: {full_coverage}, Partial: {partial_coverage}, None: {no_coverage}")
print(f"Unique 3P TCs mapped: {len(mapped_3p_ids)}")

# Write Excel
print(f"\nWriting {OUTPUT_FILE}...")

with pd.ExcelWriter(OUTPUT_FILE, engine='xlsxwriter') as writer:
    # --- TRACEABILITY SHEET ---
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Traceability", index=False)

    workbook = writer.book
    ws = writer.sheets["Traceability"]

    # Formats
    header_fmt = workbook.add_format({
        'bold': True, 'bg_color': '#1F4E79', 'font_color': 'white',
        'border': 1, 'text_wrap': True, 'valign': 'top'
    })
    high_fmt = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'bold': True})
    med_fmt = workbook.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C5700'})
    low_fmt = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    no_fmt = workbook.add_format({'bg_color': '#D9D9D9', 'font_color': '#404040', 'bold': True})
    yes_fmt = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'bold': True})
    partial_fmt = workbook.add_format({'bg_color': '#BDD7EE', 'font_color': '#1F4E79'})

    # Headers
    for col_num, col_name in enumerate(df.columns):
        ws.write(0, col_num, col_name, header_fmt)

    # Color code
    conf_col = list(df.columns).index("Confidence")
    verdict_col = list(df.columns).index("LLM Verdict")
    for row_idx in range(len(df)):
        conf = df.iloc[row_idx]["Confidence"]
        verdict = df.iloc[row_idx]["LLM Verdict"]
        if conf == "HIGH_CONFIDENCE":
            ws.write(row_idx + 1, conf_col, conf, high_fmt)
        elif conf == "MEDIUM_CONFIDENCE":
            ws.write(row_idx + 1, conf_col, conf, med_fmt)
        elif conf == "LOW_CONFIDENCE":
            ws.write(row_idx + 1, conf_col, conf, low_fmt)
        elif conf == "NO_MATCH":
            ws.write(row_idx + 1, conf_col, conf, no_fmt)

        if verdict == "YES":
            ws.write(row_idx + 1, verdict_col, verdict, yes_fmt)
        elif verdict == "PARTIAL":
            ws.write(row_idx + 1, verdict_col, verdict, partial_fmt)
        elif verdict == "NO_MATCH":
            ws.write(row_idx + 1, verdict_col, verdict, no_fmt)

    # Column widths
    col_widths = [14, 22, 40, 50, 32, 25, 14, 40, 40, 50, 50, 20, 12, 50]
    for i, w in enumerate(col_widths):
        ws.set_column(i, i, w)
    ws.freeze_panes(1, 4)

    # --- EXECUTIVE SUMMARY SHEET ---
    ws_sum = workbook.add_worksheet("Executive Summary")
    writer.sheets["Executive Summary"] = ws_sum

    title_fmt = workbook.add_format({'bold': True, 'font_size': 18, 'bottom': 2})
    subtitle_fmt = workbook.add_format({'bold': True, 'font_size': 11, 'italic': True, 'font_color': '#595959'})
    section_fmt = workbook.add_format({'bold': True, 'font_size': 12, 'bg_color': '#1F4E79', 'font_color': 'white', 'border': 1})
    label_fmt = workbook.add_format({'bold': True, 'valign': 'top'})
    value_fmt = workbook.add_format({'valign': 'top'})
    big_num_fmt = workbook.add_format({'bold': True, 'font_size': 16, 'font_color': '#1F4E79'})
    good_fmt = workbook.add_format({'font_color': '#006100', 'bold': True})
    warn_fmt = workbook.add_format({'font_color': '#9C5700', 'bold': True})
    bad_fmt = workbook.add_format({'font_color': '#9C0006', 'bold': True})
    tbl_header_fmt = workbook.add_format({'bold': True, 'bg_color': '#D6DCE4', 'border': 1})
    tbl_cell_fmt = workbook.add_format({'border': 1, 'valign': 'top'})

    row = 0
    ws_sum.write(row, 0, "Test Traceability — Executive Summary (LLM-Validated)", title_fmt)
    row += 1
    ws_sum.write(row, 0, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Method: Semantic Embeddings + LLM Intent Evaluation (Claude)", subtitle_fmt)
    row += 3

    # KEY METRICS
    ws_sum.write(row, 0, "KEY METRICS", section_fmt)
    for c in range(1, 5):
        ws_sum.write(row, c, "", section_fmt)
    row += 2

    ws_sum.write(row, 0, "Total PORTAL (Automation) Test Cases:", label_fmt)
    ws_sum.write(row, 1, total_portal, big_num_fmt)
    row += 1
    ws_sum.write(row, 0, "Total 3P (Manual) Test Cases Evaluated:", label_fmt)
    ws_sum.write(row, 1, "3,105 (corpus) → 610 pairs evaluated by LLM", value_fmt)
    row += 1
    ws_sum.write(row, 0, "Unique 3P Test Cases Mapped:", label_fmt)
    ws_sum.write(row, 1, len(mapped_3p_ids), big_num_fmt)
    row += 2

    # COVERAGE
    portal_cov_pct = (full_coverage + partial_coverage) / total_portal * 100
    ws_sum.write(row, 0, "PORTAL AUTOMATION COVERAGE", section_fmt)
    for c in range(1, 5):
        ws_sum.write(row, c, "", section_fmt)
    row += 2

    ws_sum.write(row, 0, "Full Coverage (YES — same intent & outcome):", label_fmt)
    ws_sum.write(row, 1, full_coverage, good_fmt)
    ws_sum.write(row, 2, f"({full_coverage/total_portal*100:.0f}%)", good_fmt)
    row += 1
    ws_sum.write(row, 0, "Partial Coverage (PARTIAL — related but not identical):", label_fmt)
    ws_sum.write(row, 1, partial_coverage, warn_fmt)
    ws_sum.write(row, 2, f"({partial_coverage/total_portal*100:.0f}%)", warn_fmt)
    row += 1
    ws_sum.write(row, 0, "No Coverage (no functional equivalent found):", label_fmt)
    ws_sum.write(row, 1, no_coverage, bad_fmt)
    ws_sum.write(row, 2, f"({no_coverage/total_portal*100:.0f}%)", bad_fmt)
    row += 1
    ws_sum.write(row, 0, "Overall PORTAL TCs with at least partial coverage:", label_fmt)
    ws_sum.write(row, 1, f"{portal_cov_pct:.1f}%", big_num_fmt)
    row += 3

    # CATEGORY TABLE
    ws_sum.write(row, 0, "COVERAGE BY CATEGORY", section_fmt)
    for c in range(1, 6):
        ws_sum.write(row, c, "", section_fmt)
    row += 2

    headers = ["Category", "Total", "Full (YES)", "Partial", "No Match", "Coverage"]
    for c, h in enumerate(headers):
        ws_sum.write(row, c, h, tbl_header_fmt)
    row += 1

    for cat in sorted(category_stats.keys()):
        info = category_stats[cat]
        cov = (info["full"] + info["partial"]) / info["total"] * 100 if info["total"] > 0 else 0
        ws_sum.write(row, 0, cat, tbl_cell_fmt)
        ws_sum.write(row, 1, info["total"], tbl_cell_fmt)
        ws_sum.write(row, 2, info["full"], tbl_cell_fmt)
        ws_sum.write(row, 3, info["partial"], tbl_cell_fmt)
        ws_sum.write(row, 4, info["none"], tbl_cell_fmt)
        ws_sum.write(row, 5, f"{cov:.0f}%", tbl_cell_fmt)
        row += 1
    row += 2

    # VERDICT DISTRIBUTION
    ws_sum.write(row, 0, "LLM VERDICT DISTRIBUTION (610 pairs evaluated)", section_fmt)
    for c in range(1, 5):
        ws_sum.write(row, c, "", section_fmt)
    row += 2

    total_yes = sum(sum(1 for v in vals if v["verdict"] == "YES") for vals in all_verdicts.values())
    total_partial = sum(sum(1 for v in vals if v["verdict"] == "PARTIAL") for vals in all_verdicts.values())
    total_no = sum(sum(1 for v in vals if v["verdict"] == "NO") for vals in all_verdicts.values())

    ws_sum.write(row, 0, "YES (functionally equivalent):", label_fmt)
    ws_sum.write(row, 1, total_yes, good_fmt)
    ws_sum.write(row, 2, f"({total_yes/610*100:.1f}%)", value_fmt)
    row += 1
    ws_sum.write(row, 0, "PARTIAL (related functionality):", label_fmt)
    ws_sum.write(row, 1, total_partial, warn_fmt)
    ws_sum.write(row, 2, f"({total_partial/610*100:.1f}%)", value_fmt)
    row += 1
    ws_sum.write(row, 0, "NO (different intent):", label_fmt)
    ws_sum.write(row, 1, total_no, bad_fmt)
    ws_sum.write(row, 2, f"({total_no/610*100:.1f}%)", value_fmt)
    row += 3

    # GAPS
    ws_sum.write(row, 0, "COVERAGE GAPS — PORTAL TCs WITH NO 3P EQUIVALENT", section_fmt)
    for c in range(1, 5):
        ws_sum.write(row, c, "", section_fmt)
    row += 2

    no_cov_list = [pid for pid, cov in coverage_stats.items() if cov == "NONE"]
    if no_cov_list:
        for pid in no_cov_list:
            portal_info = eval_pairs[pid]["portal"]
            ws_sum.write(row, 0, f"• {pid}", bad_fmt)
            ws_sum.write(row, 1, portal_info["title"], value_fmt)
            ws_sum.write(row, 2, portal_info["category"], value_fmt)
            row += 1
        row += 1
        ws_sum.write(row, 0, "Recommendation:", label_fmt)
        ws_sum.write(row, 1, "Create new 3P manual test cases for these, or confirm they are automation-only scope.", value_fmt)
    else:
        ws_sum.write(row, 0, "All PORTAL TCs have at least partial coverage.", good_fmt)
    row += 3

    # PARTIAL-ONLY TCs
    partial_only = [pid for pid, cov in coverage_stats.items() if cov == "PARTIAL"]
    if partial_only:
        ws_sum.write(row, 0, "PORTAL TCs WITH ONLY PARTIAL COVERAGE (no exact match)", section_fmt)
        for c in range(1, 5):
            ws_sum.write(row, c, "", section_fmt)
        row += 2
        for pid in partial_only:
            portal_info = eval_pairs[pid]["portal"]
            ws_sum.write(row, 0, f"• {pid}", warn_fmt)
            ws_sum.write(row, 1, portal_info["title"], value_fmt)
            ws_sum.write(row, 2, portal_info["category"], value_fmt)
            row += 1
        row += 1
        ws_sum.write(row, 0, "Recommendation:", label_fmt)
        ws_sum.write(row, 1, "Review PARTIAL matches — some may be sufficient, others may need new 3P test cases.", value_fmt)

    # Column widths
    ws_sum.set_column(0, 0, 50)
    ws_sum.set_column(1, 1, 50)
    ws_sum.set_column(2, 2, 20)
    ws_sum.set_column(3, 3, 15)
    ws_sum.set_column(4, 4, 15)
    ws_sum.set_column(5, 5, 12)

print(f"Done. Output: {OUTPUT_FILE}")
