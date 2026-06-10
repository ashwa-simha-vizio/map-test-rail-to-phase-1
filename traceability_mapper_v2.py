"""
Traceability Mapper v2 — Semantic Embedding Approach
Uses sentence-transformers (all-MiniLM-L6-v2) for intent-based matching.
Upgrades over v1:
  - Semantic embeddings capture synonyms and intent, not just word overlap
  - Stricter thresholds calibrated for cosine similarity of embeddings
  - Capped at top 5 matches per PORTAL TC (quality over quantity)
  - "Best Match" flag for the single strongest mapping
  - Category compatibility as hard filter (not just a boost)
  - Manual Review flag for borderline cases
"""

import openpyxl
import pandas as pd
import re
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# ============================================================
# CONFIGURATION
# ============================================================
PORTAL_FILE = "npi_certification_sanity_report.xlsx"
PORTAL_SHEET = "Phase 1 Test Cases"

THREE_P_FILE = "vizio_3p_test_suites.xlsx"
IGNORE_SHEETS = ["Sheet1", "Report Overview", "Summary"]

OUTPUT_FILE = "traceability_output_v2.xlsx"

# Embedding model — good balance of speed and quality
MODEL_NAME = "all-MiniLM-L6-v2"

# Confidence thresholds (calibrated for sentence embeddings)
HIGH_CONFIDENCE_THRESHOLD = 0.65
MEDIUM_CONFIDENCE_THRESHOLD = 0.50
LOW_CONFIDENCE_THRESHOLD = 0.38

# Max matches per PORTAL TC
MAX_MATCHES_PER_TC = 5

# Category compatibility map — defines which PORTAL categories can map to which 3P categories/sheets
CATEGORY_COMPATIBILITY = {
    "power management": {
        "keywords": ["power", "standby", "wake", "boot", "reboot", "dc power", "ac power", "sleep", "energy"],
        "sheets": ["TV 101 Full Sweep"],
    },
    "network/connectivity": {
        "keywords": ["wifi", "wi-fi", "network", "ethernet", "bluetooth", "bt", "connection", "internet", "wireless", "5ghz", "2.4ghz", "router"],
        "sheets": ["TV 101 Full Sweep", "Wi-Fi Performance Report (SKU S", "Wi-Fi Sanity Test (SKU Specific", "BT Performance Report (SKU Spec"],
    },
    "display/video output": {
        "keywords": ["display", "video", "hdmi", "resolution", "hdr", "dolby vision", "4k", "picture", "pq", "backlight", "panel"],
        "sheets": ["TV 101 Full Sweep", "PQ Functionality (SKU Specific)"],
    },
    "audio output": {
        "keywords": ["audio", "sound", "volume", "speaker", "soundbar", "dolby", "dts", "earc", "arc", "atmos", "pcm", "stereo"],
        "sheets": ["TV 101 Full Sweep"],
    },
    "remote control (ir/bt)": {
        "keywords": ["remote", "ir", "bt", "key", "button", "control", "pair", "unpair", "voice"],
        "sheets": ["TV 101 Full Sweep", "Remote Control (IRBT) (Remote S"],
    },
    "input/hdmi/cec": {
        "keywords": ["input", "hdmi", "cec", "source", "arc", "earc", "one touch", "routing"],
        "sheets": ["TV 101 Full Sweep"],
    },
    "antenna/tuner": {
        "keywords": ["antenna", "tuner", "channel", "scan", "atsc", "broadcast", "coax", "tv tuner"],
        "sheets": ["TV 101 Full Sweep"],
    },
    "device provisioning": {
        "keywords": ["provision", "mac address", "ulpk", "serial", "factory", "calibration"],
        "sheets": ["TV 101 Full Sweep", "Mac Address ULPK Provisioning ("],
    },
    "oob/ota updates": {
        "keywords": ["ota", "upgrade", "update", "firmware", "software", "download", "oob"],
        "sheets": ["TV 101 Full Sweep", "Upgradability (OTA) Test", "OOB Test (SKU Specific)"],
    },
    "oobe/initial setup": {
        "keywords": ["oobe", "oob", "setup", "first boot", "out of box", "wizard", "initial", "welcome"],
        "sheets": ["TV 101 Full Sweep", "OOB Test (SKU Specific)"],
    },
    "streaming apps": {
        "keywords": ["app", "smartcast", "watchfree", "cast", "airplay", "homekit", "netflix", "youtube", "disney", "streaming", "launch"],
        "sheets": ["TV 101 Full Sweep"],
    },
    "settings/system ui": {
        "keywords": ["settings", "menu", "configuration", "options", "system", "ui", "parental", "closed caption", "accessibility"],
        "sheets": ["TV 101 Full Sweep", "Retail Demo (SKU Specific)"],
    },
}


def normalize_text(text):
    """Clean and normalize text for comparison."""
    if not text or str(text).strip().lower() == 'none':
        return ""
    text = str(text).lower()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def get_compatible_categories(portal_category):
    """Return the set of keywords and allowed sheets for a PORTAL category."""
    portal_cat_lower = portal_category.lower().strip()
    for key, config in CATEGORY_COMPATIBILITY.items():
        if key == portal_cat_lower or portal_cat_lower in key or key in portal_cat_lower:
            return config
    # Fallback: allow all
    return None


def is_category_compatible(portal_category, three_p_category, three_p_sheet):
    """Check if a 3P test case is category-compatible with a PORTAL test case."""
    config = get_compatible_categories(portal_category)
    if config is None:
        return True  # No restriction, allow all

    three_p_cat_lower = normalize_text(three_p_category)
    three_p_sheet_lower = three_p_sheet.lower().strip()

    # Check sheet compatibility
    sheet_ok = any(s.lower() in three_p_sheet_lower or three_p_sheet_lower in s.lower()
                   for s in config["sheets"])

    # Check keyword compatibility in category
    keyword_ok = any(kw in three_p_cat_lower for kw in config["keywords"])

    # Accept if either sheet or keyword matches
    return sheet_ok or keyword_ok


def build_portal_text(row):
    """Build semantic text representation for PORTAL test case.
    Focuses on: WHAT is being tested + HOW + EXPECTED OUTCOME
    """
    parts = []

    # Intent statement
    title = row.get("title", "")
    if title:
        parts.append(f"Test: {title}")

    # Category context
    category = row.get("category", "")
    if category:
        parts.append(f"Category: {category}")

    # Preconditions
    precond = row.get("preconditions", "")
    if precond:
        parts.append(f"Preconditions: {precond}")

    # Steps/Description
    desc = row.get("description", "")
    if desc:
        parts.append(f"Steps: {desc}")

    # Detailed description (often more specific steps)
    detailed = row.get("detailed_description", "")
    if detailed:
        parts.append(f"Detailed steps: {detailed}")

    # Expected outcome (critical for intent matching)
    expected = row.get("expected_output", "")
    if expected:
        # Trim very long expected outputs to key parts
        if len(expected) > 500:
            expected = expected[:500]
        parts.append(f"Expected outcome: {expected}")

    detailed_exp = row.get("detailed_expected_output", "")
    if detailed_exp:
        parts.append(f"Expected result: {detailed_exp}")

    return " ".join(parts)


def build_3p_text(row):
    """Build semantic text representation for 3P test case."""
    parts = []

    title = row.get("title", "")
    if title:
        parts.append(f"Test: {title}")

    category = row.get("category", "")
    if category:
        parts.append(f"Category: {category}")

    precond = row.get("preconditions", "")
    if precond:
        parts.append(f"Preconditions: {precond}")

    steps = row.get("steps", "")
    if steps:
        parts.append(f"Steps: {steps}")

    steps_sep = row.get("steps_separated", "")
    if steps_sep:
        # Steps separated contains step + expected result pairs
        if len(steps_sep) > 800:
            steps_sep = steps_sep[:800]
        parts.append(f"Step details: {steps_sep}")

    return " ".join(parts)


def load_portal_test_cases():
    """Load PORTAL (automation) test cases from Phase 1 sheet."""
    wb = openpyxl.load_workbook(PORTAL_FILE, read_only=True)
    ws = wb[PORTAL_SHEET]

    test_cases = []
    for row in ws.iter_rows(min_row=2, max_col=14, values_only=True):
        if not row[0]:
            continue
        tc = {
            "id": str(row[0]).strip() if row[0] else "",
            "category": str(row[1]).strip() if row[1] else "",
            "title": str(row[2]).strip() if row[2] else "",
            "priority": str(row[3]).strip() if row[3] else "",
            "description": str(row[5]).strip() if len(row) > 5 and row[5] else "",
            "expected_output": str(row[6]).strip() if len(row) > 6 and row[6] else "",
            "preconditions": str(row[11]).strip() if len(row) > 11 and row[11] else "",
            "detailed_description": str(row[12]).strip() if len(row) > 12 and row[12] else "",
            "detailed_expected_output": str(row[13]).strip() if len(row) > 13 and row[13] else "",
        }
        test_cases.append(tc)

    wb.close()
    print(f"Loaded {len(test_cases)} PORTAL test cases")
    return test_cases


def load_3p_test_cases():
    """Load all 3P manual test cases from relevant sheets."""
    wb = openpyxl.load_workbook(THREE_P_FILE, read_only=True)
    all_test_cases = []

    for sheet_name in wb.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue

        ws = wb[sheet_name]

        # Find the header row
        header_row = None
        header_map = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            row_vals = [str(v).strip().lower() if v else "" for v in row]
            if "id" in row_vals and "title" in row_vals:
                header_row = i
                for j, val in enumerate(row_vals):
                    if val:
                        header_map[val] = j
                break

        if header_row is None:
            print(f"  Skipping sheet '{sheet_name}' - no header row found")
            continue

        id_col = header_map.get("id")
        title_col = header_map.get("title")
        category_col = header_map.get("category")
        precond_col = header_map.get("preconds")
        steps_col = header_map.get("steps")
        steps_sep_col = header_map.get("steps separated")

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_list = list(row)
            tc_id = row_list[id_col] if id_col is not None and id_col < len(row_list) else None
            if not tc_id:
                continue

            tc = {
                "sheet": sheet_name,
                "id": str(tc_id).strip(),
                "title": str(row_list[title_col]).strip() if title_col is not None and title_col < len(row_list) and row_list[title_col] else "",
                "category": str(row_list[category_col]).strip() if category_col is not None and category_col < len(row_list) and row_list[category_col] else "",
                "preconditions": str(row_list[precond_col]).strip() if precond_col is not None and precond_col < len(row_list) and row_list[precond_col] else "",
                "steps": str(row_list[steps_col]).strip() if steps_col is not None and steps_col < len(row_list) and row_list[steps_col] else "",
                "steps_separated": str(row_list[steps_sep_col]).strip() if steps_sep_col is not None and steps_sep_col < len(row_list) and row_list[steps_sep_col] else "",
            }
            all_test_cases.append(tc)
            count += 1

        print(f"  Sheet '{sheet_name}': {count} test cases")

    wb.close()
    print(f"Loaded {len(all_test_cases)} total 3P test cases")
    return all_test_cases


def compute_mappings(portal_tcs, three_p_tcs, model):
    """Compute traceability mappings using semantic embeddings."""
    print("\nBuilding text representations...")

    portal_texts = [build_portal_text(tc) for tc in portal_tcs]
    three_p_texts = [build_3p_text(tc) for tc in three_p_tcs]

    print(f"Encoding {len(portal_texts)} PORTAL texts...")
    portal_embeddings = model.encode(portal_texts, show_progress_bar=True, batch_size=32)

    print(f"Encoding {len(three_p_texts)} 3P texts...")
    three_p_embeddings = model.encode(three_p_texts, show_progress_bar=True, batch_size=64)

    print("Computing similarity matrix...")
    sim_matrix = cosine_similarity(portal_embeddings, three_p_embeddings)

    print("Mapping test cases with category filtering...\n")
    mappings = []

    for i, portal_tc in enumerate(portal_tcs):
        similarities = sim_matrix[i]

        # Apply category compatibility filter
        compatible_indices = []
        for j, three_p_tc in enumerate(three_p_tcs):
            if is_category_compatible(portal_tc["category"], three_p_tc["category"], three_p_tc["sheet"]):
                compatible_indices.append(j)

        if not compatible_indices:
            # Fallback: use all if no compatible found
            compatible_indices = list(range(len(three_p_tcs)))

        # Get scores for compatible candidates only
        compatible_scores = [(j, similarities[j]) for j in compatible_indices]
        compatible_scores.sort(key=lambda x: x[1], reverse=True)

        # Select top matches above threshold
        matched = False
        match_count = 0
        best_score = 0

        for rank, (idx, score) in enumerate(compatible_scores):
            if match_count >= MAX_MATCHES_PER_TC:
                break
            if score < LOW_CONFIDENCE_THRESHOLD:
                break

            # Determine confidence
            if score >= HIGH_CONFIDENCE_THRESHOLD:
                confidence = "HIGH_CONFIDENCE"
            elif score >= MEDIUM_CONFIDENCE_THRESHOLD:
                confidence = "MEDIUM_CONFIDENCE"
            else:
                confidence = "LOW_CONFIDENCE"

            three_p_tc = three_p_tcs[idx]
            is_best = (match_count == 0)
            if is_best:
                best_score = score

            mappings.append({
                "portal_id": portal_tc["id"],
                "portal_category": portal_tc["category"],
                "portal_title": portal_tc["title"],
                "portal_description": portal_tc["description"],
                "three_p_sheet": three_p_tc["sheet"],
                "three_p_category": three_p_tc["category"],
                "three_p_id": three_p_tc["id"],
                "three_p_title": three_p_tc["title"],
                "three_p_preconditions": three_p_tc["preconditions"],
                "three_p_steps": three_p_tc["steps"],
                "three_p_steps_separated": three_p_tc["steps_separated"],
                "confidence": confidence,
                "similarity_score": round(float(score), 4),
                "best_match": "YES" if is_best else "",
                "needs_review": "YES" if 0.38 <= score <= 0.52 else "",
            })
            matched = True
            match_count += 1

        if not matched:
            mappings.append({
                "portal_id": portal_tc["id"],
                "portal_category": portal_tc["category"],
                "portal_title": portal_tc["title"],
                "portal_description": portal_tc["description"],
                "three_p_sheet": "",
                "three_p_category": "",
                "three_p_id": "",
                "three_p_title": "",
                "three_p_preconditions": "",
                "three_p_steps": "",
                "three_p_steps_separated": "",
                "confidence": "NO_MATCH",
                "similarity_score": 0.0,
                "best_match": "",
                "needs_review": "YES",
            })

        # Progress
        conf_str = "NO_MATCH" if not matched else f"top={best_score:.3f}"
        print(f"  {portal_tc['id']:12s} | {portal_tc['category']:25s} | {match_count} matches | {conf_str}")

    return mappings


def generate_executive_summary(mappings, portal_tcs, three_p_tcs):
    """Generate executive summary statistics."""
    total_portal = len(portal_tcs)

    # Group by portal TC
    portal_coverage = {}
    for tc in portal_tcs:
        portal_coverage[tc["id"]] = {
            "category": tc["category"],
            "confidences": [],
            "best_score": 0,
            "match_count": 0,
        }

    for m in mappings:
        pid = m["portal_id"]
        if pid in portal_coverage and m["confidence"] != "NO_MATCH":
            portal_coverage[pid]["confidences"].append(m["confidence"])
            portal_coverage[pid]["best_score"] = max(portal_coverage[pid]["best_score"], m["similarity_score"])
            portal_coverage[pid]["match_count"] += 1

    # Coverage breakdown
    full_coverage = 0
    partial_coverage = 0
    no_coverage = 0

    for pid, info in portal_coverage.items():
        if not info["confidences"]:
            no_coverage += 1
        elif "HIGH_CONFIDENCE" in info["confidences"]:
            full_coverage += 1
        else:
            partial_coverage += 1

    # Unique 3P IDs mapped
    mapped_3p_ids = set()
    for m in mappings:
        if m["confidence"] != "NO_MATCH" and m["three_p_id"]:
            mapped_3p_ids.add(m["three_p_id"])

    total_3p_mapped = len(mapped_3p_ids)
    total_3p = len(three_p_tcs)

    # Primary metric: PORTAL coverage (what % of automation TCs have manual equivalents)
    portal_coverage_pct = ((full_coverage + partial_coverage) / total_portal * 100) if total_portal > 0 else 0

    # Secondary metric: 3P corpus coverage
    three_p_coverage_pct = (total_3p_mapped / total_3p * 100) if total_3p > 0 else 0

    # Category breakdown
    category_stats = {}
    for pid, info in portal_coverage.items():
        cat = info["category"]
        if cat not in category_stats:
            category_stats[cat] = {"total": 0, "full": 0, "partial": 0, "none": 0}
        category_stats[cat]["total"] += 1
        if "HIGH_CONFIDENCE" in info["confidences"]:
            category_stats[cat]["full"] += 1
        elif info["confidences"]:
            category_stats[cat]["partial"] += 1
        else:
            category_stats[cat]["none"] += 1

    # 3P suite distribution
    suite_mapping_count = {}
    for m in mappings:
        if m["confidence"] != "NO_MATCH" and m["three_p_sheet"]:
            sheet = m["three_p_sheet"]
            if sheet not in suite_mapping_count:
                suite_mapping_count[sheet] = 0
            suite_mapping_count[sheet] += 1

    # Confidence distribution
    conf_dist = {"HIGH_CONFIDENCE": 0, "MEDIUM_CONFIDENCE": 0, "LOW_CONFIDENCE": 0, "NO_MATCH": 0}
    for m in mappings:
        conf_dist[m["confidence"]] += 1

    return {
        "total_portal": total_portal,
        "total_3p": total_3p,
        "total_3p_mapped": total_3p_mapped,
        "portal_coverage_pct": portal_coverage_pct,
        "three_p_coverage_pct": three_p_coverage_pct,
        "full_coverage": full_coverage,
        "partial_coverage": partial_coverage,
        "no_coverage": no_coverage,
        "category_stats": category_stats,
        "suite_mapping_count": suite_mapping_count,
        "conf_dist": conf_dist,
        "portal_coverage": portal_coverage,
    }


def write_output(mappings, summary, portal_tcs, three_p_tcs):
    """Write formatted Excel output."""
    print(f"\nWriting output to {OUTPUT_FILE}...")

    with pd.ExcelWriter(OUTPUT_FILE, engine='xlsxwriter') as writer:
        # ---- TRACEABILITY SHEET ----
        trace_df = pd.DataFrame(mappings)
        trace_df.columns = [
            "PORTAL Test Case ID",
            "PORTAL Category",
            "PORTAL Title",
            "PORTAL Description",
            "3P Sheet/Suite",
            "3P Category",
            "3P Test Case ID",
            "3P Title",
            "3P Preconditions",
            "3P Steps",
            "3P Steps Separated",
            "Confidence",
            "Similarity Score",
            "Best Match",
            "Needs Review",
        ]
        trace_df.to_excel(writer, sheet_name="Traceability", index=False)

        workbook = writer.book
        ws_trace = writer.sheets["Traceability"]

        # Formats
        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#2F5496', 'font_color': 'white',
            'border': 1, 'text_wrap': True, 'valign': 'top'
        })
        high_fmt = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'bold': True})
        med_fmt = workbook.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C5700'})
        low_fmt = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        no_fmt = workbook.add_format({'bg_color': '#D9D9D9', 'font_color': '#404040', 'bold': True})
        best_fmt = workbook.add_format({'bg_color': '#BDD7EE', 'bold': True})
        review_fmt = workbook.add_format({'bg_color': '#FCE4D6', 'font_color': '#833C0B'})

        # Write headers
        for col_num, col_name in enumerate(trace_df.columns):
            ws_trace.write(0, col_num, col_name, header_fmt)

        # Color-code confidence and flags
        conf_col = trace_df.columns.get_loc("Confidence")
        best_col = trace_df.columns.get_loc("Best Match")
        review_col = trace_df.columns.get_loc("Needs Review")

        for row_idx in range(len(trace_df)):
            conf = trace_df.iloc[row_idx]["Confidence"]
            if conf == "HIGH_CONFIDENCE":
                ws_trace.write(row_idx + 1, conf_col, conf, high_fmt)
            elif conf == "MEDIUM_CONFIDENCE":
                ws_trace.write(row_idx + 1, conf_col, conf, med_fmt)
            elif conf == "LOW_CONFIDENCE":
                ws_trace.write(row_idx + 1, conf_col, conf, low_fmt)
            elif conf == "NO_MATCH":
                ws_trace.write(row_idx + 1, conf_col, conf, no_fmt)

            if trace_df.iloc[row_idx]["Best Match"] == "YES":
                ws_trace.write(row_idx + 1, best_col, "YES", best_fmt)
            if trace_df.iloc[row_idx]["Needs Review"] == "YES":
                ws_trace.write(row_idx + 1, review_col, "YES", review_fmt)

        # Column widths
        col_widths = [14, 22, 40, 50, 32, 25, 14, 40, 40, 50, 50, 20, 12, 12, 13]
        for i, w in enumerate(col_widths):
            ws_trace.set_column(i, i, w)

        # Freeze top row and first 4 columns
        ws_trace.freeze_panes(1, 4)

        # ---- EXECUTIVE SUMMARY SHEET ----
        ws_summary = workbook.add_worksheet("Executive Summary")
        writer.sheets["Executive Summary"] = ws_summary

        # Formats
        title_fmt = workbook.add_format({'bold': True, 'font_size': 18, 'bottom': 2})
        subtitle_fmt = workbook.add_format({'bold': True, 'font_size': 11, 'italic': True, 'font_color': '#595959'})
        section_fmt = workbook.add_format({'bold': True, 'font_size': 12, 'bg_color': '#2F5496', 'font_color': 'white', 'border': 1})
        label_fmt = workbook.add_format({'bold': True, 'valign': 'top'})
        value_fmt = workbook.add_format({'valign': 'top'})
        big_num_fmt = workbook.add_format({'bold': True, 'font_size': 16, 'font_color': '#2F5496'})
        pct_fmt = workbook.add_format({'bold': True, 'font_size': 16, 'font_color': '#2F5496', 'num_format': '0.0"%"'})
        good_fmt = workbook.add_format({'font_color': '#006100', 'bold': True})
        warn_fmt = workbook.add_format({'font_color': '#9C5700', 'bold': True})
        bad_fmt = workbook.add_format({'font_color': '#9C0006', 'bold': True})
        tbl_header_fmt = workbook.add_format({'bold': True, 'bg_color': '#D6DCE4', 'border': 1})
        tbl_cell_fmt = workbook.add_format({'border': 1, 'valign': 'top'})

        row = 0
        ws_summary.write(row, 0, "Automation Test Traceability — Executive Summary", title_fmt)
        row += 1
        ws_summary.write(row, 0, f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')} | Model: {MODEL_NAME} | Method: Semantic Embedding Similarity", subtitle_fmt)
        row += 3

        # === KEY METRICS ===
        ws_summary.write(row, 0, "KEY METRICS", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        ws_summary.write(row, 3, "", section_fmt)
        row += 2

        ws_summary.write(row, 0, "PORTAL Automation Test Cases", label_fmt)
        ws_summary.write(row, 1, summary["total_portal"], big_num_fmt)
        ws_summary.write(row, 2, "3P Manual Test Cases (corpus)", label_fmt)
        ws_summary.write(row, 3, summary["total_3p"], big_num_fmt)
        row += 2

        ws_summary.write(row, 0, "PORTAL Test Cases with Manual Coverage", label_fmt)
        ws_summary.write(row, 1, f"{summary['portal_coverage_pct']:.1f}%", pct_fmt)
        row += 1
        ws_summary.write(row, 0, "(% of automation TCs that map to at least one 3P manual TC)", subtitle_fmt)
        row += 2

        ws_summary.write(row, 0, "Unique 3P Test Cases Referenced", label_fmt)
        ws_summary.write(row, 1, summary["total_3p_mapped"], big_num_fmt)
        ws_summary.write(row, 2, f"({summary['three_p_coverage_pct']:.1f}% of 3P corpus)", value_fmt)
        row += 3

        # === COVERAGE BREAKDOWN ===
        ws_summary.write(row, 0, "PORTAL COVERAGE BREAKDOWN", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        ws_summary.write(row, 3, "", section_fmt)
        row += 2

        ws_summary.write(row, 0, "Full Coverage (≥1 HIGH_CONFIDENCE match):", label_fmt)
        ws_summary.write(row, 1, summary["full_coverage"], good_fmt)
        row += 1
        ws_summary.write(row, 0, "Partial Coverage (MEDIUM/LOW only):", label_fmt)
        ws_summary.write(row, 1, summary["partial_coverage"], warn_fmt)
        row += 1
        ws_summary.write(row, 0, "No Coverage (NO_MATCH):", label_fmt)
        ws_summary.write(row, 1, summary["no_coverage"], bad_fmt)
        row += 2

        # Confidence distribution
        ws_summary.write(row, 0, "Mapping Confidence Distribution:", label_fmt)
        row += 1
        for conf, count in summary["conf_dist"].items():
            ws_summary.write(row, 0, f"  {conf}", value_fmt)
            ws_summary.write(row, 1, count, value_fmt)
            row += 1
        row += 2

        # === CATEGORY COVERAGE TABLE ===
        ws_summary.write(row, 0, "COVERAGE BY PORTAL CATEGORY", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        ws_summary.write(row, 3, "", section_fmt)
        ws_summary.write(row, 4, "", section_fmt)
        row += 2

        headers = ["Category", "Total TCs", "Full (HIGH)", "Partial (MED/LOW)", "No Match"]
        for c, h in enumerate(headers):
            ws_summary.write(row, c, h, tbl_header_fmt)
        row += 1

        for cat in sorted(summary["category_stats"].keys()):
            info = summary["category_stats"][cat]
            ws_summary.write(row, 0, cat, tbl_cell_fmt)
            ws_summary.write(row, 1, info["total"], tbl_cell_fmt)
            ws_summary.write(row, 2, info["full"], tbl_cell_fmt)
            ws_summary.write(row, 3, info["partial"], tbl_cell_fmt)
            ws_summary.write(row, 4, info["none"], tbl_cell_fmt)
            row += 1
        row += 2

        # === 3P SUITE DISTRIBUTION ===
        ws_summary.write(row, 0, "3P SUITE MAPPING DISTRIBUTION", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        row += 2

        ws_summary.write(row, 0, "3P Suite/Sheet", tbl_header_fmt)
        ws_summary.write(row, 1, "Mappings", tbl_header_fmt)
        row += 1

        for sheet, count in sorted(summary["suite_mapping_count"].items(), key=lambda x: x[1], reverse=True):
            ws_summary.write(row, 0, sheet, tbl_cell_fmt)
            ws_summary.write(row, 1, count, tbl_cell_fmt)
            row += 1
        row += 2

        # === GAPS & RECOMMENDATIONS ===
        ws_summary.write(row, 0, "GAPS & RECOMMENDATIONS", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        ws_summary.write(row, 3, "", section_fmt)
        row += 2

        # No-coverage test cases
        no_match_tcs = [m for m in mappings if m["confidence"] == "NO_MATCH"]
        if no_match_tcs:
            ws_summary.write(row, 0, "PORTAL Test Cases with NO manual test equivalent:", label_fmt)
            row += 1
            for m in no_match_tcs:
                ws_summary.write(row, 0, f"  • {m['portal_id']}: {m['portal_title']}", value_fmt)
                row += 1
            row += 1
            ws_summary.write(row, 0, "Action: Create corresponding 3P manual test cases or confirm automation-only scope.", subtitle_fmt)
            row += 2

        # Categories with gaps
        gap_cats = [(cat, info) for cat, info in summary["category_stats"].items()
                    if info["none"] > 0 or info["full"] == 0]
        if gap_cats:
            ws_summary.write(row, 0, "Categories needing attention:", label_fmt)
            row += 1
            for cat, info in gap_cats:
                if info["none"] > 0:
                    ws_summary.write(row, 0, f"  • {cat}: {info['none']} TC(s) with no manual coverage", bad_fmt)
                elif info["full"] == 0:
                    ws_summary.write(row, 0, f"  • {cat}: No HIGH_CONFIDENCE matches (only partial alignment)", warn_fmt)
                row += 1

        # Column widths
        ws_summary.set_column(0, 0, 48)
        ws_summary.set_column(1, 1, 20)
        ws_summary.set_column(2, 2, 20)
        ws_summary.set_column(3, 3, 20)
        ws_summary.set_column(4, 4, 15)

    print(f"Output written to: {OUTPUT_FILE}")


def main():
    print("=" * 70)
    print("  TRACEABILITY MAPPER v2 — Semantic Embedding Approach")
    print("=" * 70)

    # Load model from local cache
    print(f"\nLoading embedding model: {MODEL_NAME}...")
    import os
    os.environ["HF_HUB_OFFLINE"] = "1"
    os.environ["TRANSFORMERS_OFFLINE"] = "1"
    model = SentenceTransformer(MODEL_NAME)
    print("Model loaded.\n")

    # Load test cases
    print("--- Loading PORTAL test cases ---")
    portal_tcs = load_portal_test_cases()

    print("\n--- Loading 3P test cases ---")
    three_p_tcs = load_3p_test_cases()

    # Compute mappings
    print("\n--- Computing semantic mappings ---")
    mappings = compute_mappings(portal_tcs, three_p_tcs, model)

    # Generate summary
    print("\n--- Generating executive summary ---")
    summary = generate_executive_summary(mappings, portal_tcs, three_p_tcs)

    # Write output
    write_output(mappings, summary, portal_tcs, three_p_tcs)

    # Console summary
    print("\n" + "=" * 70)
    print("  RESULTS SUMMARY")
    print("=" * 70)
    print(f"  PORTAL test cases:          {summary['total_portal']}")
    print(f"  PORTAL with coverage:       {summary['full_coverage'] + summary['partial_coverage']} ({summary['portal_coverage_pct']:.1f}%)")
    print(f"    - Full (HIGH):            {summary['full_coverage']}")
    print(f"    - Partial (MED/LOW):      {summary['partial_coverage']}")
    print(f"    - No match:               {summary['no_coverage']}")
    print(f"  Unique 3P TCs referenced:   {summary['total_3p_mapped']} / {summary['total_3p']}")
    print(f"  Confidence distribution:")
    for conf, count in summary["conf_dist"].items():
        print(f"    - {conf:20s}: {count}")
    print(f"\n  Output: {OUTPUT_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
