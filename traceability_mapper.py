"""
Traceability Mapper: Maps PORTAL automation test cases to 3P manual test cases.
Uses TF-IDF cosine similarity + category alignment + intent matching.
"""

import openpyxl
import pandas as pd
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# ============================================================
# CONFIGURATION
# ============================================================
PORTAL_FILE = "npi_certification_sanity_report.xlsx"
PORTAL_SHEET = "Phase 1 Test Cases"

THREE_P_FILE = "vizio_3p_test_suites.xlsx"
IGNORE_SHEETS = ["Sheet1", "Report Overview", "Summary"]

OUTPUT_FILE = "traceability_output.xlsx"

# Similarity thresholds
HIGH_CONFIDENCE_THRESHOLD = 0.55
MEDIUM_CONFIDENCE_THRESHOLD = 0.38
LOW_CONFIDENCE_THRESHOLD = 0.25

# Maximum matches per PORTAL test case
MAX_MATCHES_PER_TC = 15

# ============================================================
# CATEGORY MAPPING (PORTAL categories -> likely 3P sheet/category keywords)
# ============================================================
CATEGORY_AFFINITY = {
    "power management": ["power", "standby", "wake", "boot", "reboot", "dc power", "ac power"],
    "network & connectivity": ["wifi", "wi-fi", "network", "ethernet", "bluetooth", "bt", "connection"],
    "input management": ["input", "hdmi", "cec", "source", "airplay", "cast"],
    "audio": ["audio", "sound", "volume", "speaker", "soundbar", "dolby", "dts", "earc", "arc"],
    "remote control": ["remote", "ir", "bt", "key", "button", "control"],
    "ota & updates": ["ota", "upgrade", "update", "firmware", "software"],
    "display & picture": ["pq", "picture", "display", "resolution", "hdr", "backlight"],
    "apps & smartcast": ["app", "smartcast", "watchfree", "cast", "airplay", "homekit"],
    "oobe": ["oob", "oobe", "setup", "first boot", "out of box"],
    "settings": ["settings", "menu", "configuration", "options"],
    "usb & media": ["usb", "media", "photo", "video", "music"],
}


def normalize_text(text):
    """Clean and normalize text for comparison."""
    if not text:
        return ""
    text = str(text).lower()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def build_portal_composite(row):
    """Build a composite text from PORTAL test case for similarity matching."""
    parts = []
    # Title (high weight - repeat)
    title = normalize_text(row.get("title", ""))
    parts.append(title)
    parts.append(title)  # double weight

    # Description (Column F - the steps)
    desc = normalize_text(row.get("description", ""))
    parts.append(desc)

    # Expected Output
    expected = normalize_text(row.get("expected_output", ""))
    parts.append(expected)

    # Pre-conditions
    precond = normalize_text(row.get("preconditions", ""))
    parts.append(precond)

    # Detailed Description
    detailed = normalize_text(row.get("detailed_description", ""))
    parts.append(detailed)

    # Detailed Expected Output
    detailed_exp = normalize_text(row.get("detailed_expected_output", ""))
    parts.append(detailed_exp)

    return " ".join(parts)


def build_3p_composite(row):
    """Build a composite text from 3P test case for similarity matching."""
    parts = []
    # Title (high weight)
    title = normalize_text(row.get("title", ""))
    parts.append(title)
    parts.append(title)  # double weight

    # Preconditions
    precond = normalize_text(row.get("preconditions", ""))
    parts.append(precond)

    # Steps
    steps = normalize_text(row.get("steps", ""))
    parts.append(steps)

    # Steps Separated
    steps_sep = normalize_text(row.get("steps_separated", ""))
    parts.append(steps_sep)

    # Category
    cat = normalize_text(row.get("category", ""))
    parts.append(cat)

    return " ".join(parts)


def category_boost(portal_category, three_p_category, three_p_sheet):
    """Calculate a category alignment boost score."""
    portal_cat_lower = normalize_text(portal_category)
    three_p_cat_lower = normalize_text(three_p_category)
    three_p_sheet_lower = normalize_text(three_p_sheet)

    # Direct category match
    if portal_cat_lower and three_p_cat_lower:
        if portal_cat_lower in three_p_cat_lower or three_p_cat_lower in portal_cat_lower:
            return 0.15

    # Check affinity mapping
    for portal_key, keywords in CATEGORY_AFFINITY.items():
        if any(kw in portal_cat_lower for kw in keywords) or portal_cat_lower in portal_key:
            # Check if 3P matches these keywords
            combined_3p = three_p_cat_lower + " " + three_p_sheet_lower
            if any(kw in combined_3p for kw in keywords):
                return 0.10

    return 0.0


def load_portal_test_cases():
    """Load PORTAL (automation) test cases from Phase 1 sheet."""
    wb = openpyxl.load_workbook(PORTAL_FILE, read_only=True)
    ws = wb[PORTAL_SHEET]

    test_cases = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=14, values_only=True)):
        if not row[0]:  # Skip empty rows
            continue
        tc = {
            "id": str(row[0]).strip() if row[0] else "",
            "category": str(row[1]).strip() if row[1] else "",
            "title": str(row[2]).strip() if row[2] else "",
            "priority": str(row[3]).strip() if row[3] else "",
            "description": str(row[5]).strip() if row[5] else "",
            "expected_output": str(row[6]).strip() if row[6] else "",
            "preconditions": str(row[11]).strip() if row[11] else "",
            "detailed_description": str(row[12]).strip() if row[12] else "",
            "detailed_expected_output": str(row[13]).strip() if row[13] else "",
        }
        test_cases.append(tc)

    wb.close()
    print(f"Loaded {len(test_cases)} PORTAL test cases")
    return test_cases


def load_3p_test_cases():
    """Load all 3P manual test cases from relevant sheets."""
    wb = openpyxl.load_workbook(THREE_P_FILE, read_only=True)
    all_test_cases = []

    for sheet_name in wb.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue

        ws = wb[sheet_name]

        # Find the header row (contains 'Id' or 'Category')
        header_row = None
        header_map = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            row_vals = [str(v).strip().lower() if v else "" for v in row]
            if "id" in row_vals and "title" in row_vals:
                header_row = i
                for j, val in enumerate(row_vals):
                    if val:
                        header_map[val] = j
                break

        if header_row is None:
            print(f"  Skipping sheet '{sheet_name}' - no header row found")
            continue

        # Map column indices
        id_col = header_map.get("id")
        title_col = header_map.get("title")
        category_col = header_map.get("category")
        precond_col = header_map.get("preconds")
        steps_col = header_map.get("steps")
        steps_sep_col = header_map.get("steps separated")

        # Read test cases
        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_list = list(row)
            tc_id = row_list[id_col] if id_col is not None and id_col < len(row_list) else None
            if not tc_id:
                continue

            tc = {
                "sheet": sheet_name,
                "id": str(tc_id).strip(),
                "title": str(row_list[title_col]).strip() if title_col is not None and title_col < len(row_list) and row_list[title_col] else "",
                "category": str(row_list[category_col]).strip() if category_col is not None and category_col < len(row_list) and row_list[category_col] else "",
                "preconditions": str(row_list[precond_col]).strip() if precond_col is not None and precond_col < len(row_list) and row_list[precond_col] else "",
                "steps": str(row_list[steps_col]).strip() if steps_col is not None and steps_col < len(row_list) and row_list[steps_col] else "",
                "steps_separated": str(row_list[steps_sep_col]).strip() if steps_sep_col is not None and steps_sep_col < len(row_list) and row_list[steps_sep_col] else "",
            }
            all_test_cases.append(tc)
            count += 1

        print(f"  Sheet '{sheet_name}': {count} test cases")

    wb.close()
    print(f"Loaded {len(all_test_cases)} total 3P test cases")
    return all_test_cases


def compute_mappings(portal_tcs, three_p_tcs):
    """Compute traceability mappings using TF-IDF + cosine similarity."""
    print("\nBuilding text representations...")

    # Build composite texts
    portal_texts = [build_portal_composite(tc) for tc in portal_tcs]
    three_p_texts = [build_3p_composite(tc) for tc in three_p_tcs]

    # Combine all texts for TF-IDF fitting
    all_texts = portal_texts + three_p_texts

    print("Computing TF-IDF vectors...")
    vectorizer = TfidfVectorizer(
        max_features=10000,
        ngram_range=(1, 2),
        stop_words='english',
        min_df=1,
        max_df=0.95
    )
    tfidf_matrix = vectorizer.fit_transform(all_texts)

    portal_vectors = tfidf_matrix[:len(portal_tcs)]
    three_p_vectors = tfidf_matrix[len(portal_tcs):]

    print("Computing cosine similarity matrix...")
    sim_matrix = cosine_similarity(portal_vectors, three_p_vectors)

    print("Mapping test cases...\n")
    mappings = []

    for i, portal_tc in enumerate(portal_tcs):
        similarities = sim_matrix[i]

        # Apply category boost
        boosted_scores = []
        for j, three_p_tc in enumerate(three_p_tcs):
            boost = category_boost(portal_tc["category"], three_p_tc["category"], three_p_tc["sheet"])
            boosted_score = similarities[j] + boost
            boosted_scores.append(boosted_score)

        boosted_scores = np.array(boosted_scores)

        # Get top matches above threshold
        top_indices = np.argsort(boosted_scores)[::-1]

        matched = False
        match_count = 0
        for idx in top_indices:
            if match_count >= MAX_MATCHES_PER_TC:
                break

            score = boosted_scores[idx]
            raw_score = similarities[idx]

            if score < LOW_CONFIDENCE_THRESHOLD:
                break

            # Determine confidence
            if score >= HIGH_CONFIDENCE_THRESHOLD:
                confidence = "HIGH_CONFIDENCE"
            elif score >= MEDIUM_CONFIDENCE_THRESHOLD:
                confidence = "MEDIUM_CONFIDENCE"
            else:
                confidence = "LOW_CONFIDENCE"

            three_p_tc = three_p_tcs[idx]
            mappings.append({
                "portal_id": portal_tc["id"],
                "portal_category": portal_tc["category"],
                "portal_title": portal_tc["title"],
                "portal_description": portal_tc["description"],
                "three_p_sheet": three_p_tc["sheet"],
                "three_p_category": three_p_tc["category"],
                "three_p_id": three_p_tc["id"],
                "three_p_title": three_p_tc["title"],
                "three_p_preconditions": three_p_tc["preconditions"],
                "three_p_steps": three_p_tc["steps"],
                "three_p_steps_separated": three_p_tc["steps_separated"],
                "confidence": confidence,
                "similarity_score": round(score, 4),
            })
            matched = True
            match_count += 1

        if not matched:
            mappings.append({
                "portal_id": portal_tc["id"],
                "portal_category": portal_tc["category"],
                "portal_title": portal_tc["title"],
                "portal_description": portal_tc["description"],
                "three_p_sheet": "",
                "three_p_category": "",
                "three_p_id": "",
                "three_p_title": "",
                "three_p_preconditions": "",
                "three_p_steps": "",
                "three_p_steps_separated": "",
                "confidence": "NO_MATCH",
                "similarity_score": 0.0,
            })

        # Progress
        top_score = boosted_scores[top_indices[0]] if len(top_indices) > 0 else 0
        print(f"  {portal_tc['id']} ({portal_tc['category']}): {match_count} matches, top score={top_score:.3f}")

    return mappings


def generate_executive_summary(mappings, portal_tcs, three_p_tcs):
    """Generate executive summary statistics."""
    total_portal = len(portal_tcs)

    # Group mappings by portal test case
    portal_coverage = {}
    for m in mappings:
        pid = m["portal_id"]
        if pid not in portal_coverage:
            portal_coverage[pid] = {"confidences": [], "categories": set()}
        if m["confidence"] != "NO_MATCH":
            portal_coverage[pid]["confidences"].append(m["confidence"])
            if m["three_p_category"]:
                portal_coverage[pid]["categories"].add(m["three_p_category"])

    # Coverage stats
    full_coverage = 0  # At least 1 HIGH match
    partial_coverage = 0  # Has MEDIUM or LOW but no HIGH
    no_coverage = 0

    for pid, info in portal_coverage.items():
        if not info["confidences"]:
            no_coverage += 1
        elif "HIGH_CONFIDENCE" in info["confidences"]:
            full_coverage += 1
        else:
            partial_coverage += 1

    # Total unique 3P mapped
    mapped_3p_ids = set()
    for m in mappings:
        if m["confidence"] != "NO_MATCH" and m["three_p_id"]:
            mapped_3p_ids.add(m["three_p_id"])

    total_3p_mapped = len(mapped_3p_ids)
    total_3p = len(three_p_tcs)
    coverage_pct = (total_3p_mapped / total_3p * 100) if total_3p > 0 else 0

    # Category coverage
    category_counts = {}
    category_portal = {}
    for m in mappings:
        pcat = m["portal_category"]
        if pcat not in category_portal:
            category_portal[pcat] = {"total": 0, "covered": 0}

    for tc in portal_tcs:
        cat = tc["category"]
        if cat not in category_portal:
            category_portal[cat] = {"total": 0, "covered": 0}
        category_portal[cat]["total"] += 1

    for pid, info in portal_coverage.items():
        # Find category for this portal TC
        for tc in portal_tcs:
            if tc["id"] == pid:
                cat = tc["category"]
                if info["confidences"]:
                    category_portal[cat]["covered"] += 1
                break

    # 3P category coverage
    three_p_cat_mapped = {}
    for m in mappings:
        if m["confidence"] != "NO_MATCH" and m["three_p_sheet"]:
            sheet = m["three_p_sheet"]
            if sheet not in three_p_cat_mapped:
                three_p_cat_mapped[sheet] = 0
            three_p_cat_mapped[sheet] += 1

    summary = {
        "total_portal": total_portal,
        "total_3p": total_3p,
        "total_3p_mapped": total_3p_mapped,
        "coverage_pct": coverage_pct,
        "full_coverage": full_coverage,
        "partial_coverage": partial_coverage,
        "no_coverage": no_coverage,
        "category_portal": category_portal,
        "three_p_cat_mapped": three_p_cat_mapped,
    }
    return summary


def write_output(mappings, summary, portal_tcs, three_p_tcs):
    """Write the output Excel workbook."""
    print("\nWriting output to", OUTPUT_FILE)

    with pd.ExcelWriter(OUTPUT_FILE, engine='xlsxwriter') as writer:
        # ---- TRACEABILITY SHEET ----
        trace_df = pd.DataFrame(mappings)
        trace_df.columns = [
            "PORTAL Test Case ID",
            "PORTAL Category",
            "PORTAL Title",
            "PORTAL Description",
            "3P Sheet/Suite",
            "3P Category",
            "3P Test Case ID",
            "3P Title",
            "3P Preconditions",
            "3P Steps",
            "3P Steps Separated",
            "Confidence",
            "Similarity Score",
        ]
        trace_df.to_excel(writer, sheet_name="Traceability", index=False)

        # Format traceability sheet
        workbook = writer.book
        ws_trace = writer.sheets["Traceability"]

        # Header format
        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'text_wrap': True, 'valign': 'top'
        })
        for col_num, col_name in enumerate(trace_df.columns):
            ws_trace.write(0, col_num, col_name, header_fmt)

        # Confidence color coding
        high_fmt = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
        med_fmt = workbook.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C5700'})
        low_fmt = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        no_fmt = workbook.add_format({'bg_color': '#D9D9D9', 'font_color': '#404040'})

        conf_col = trace_df.columns.get_loc("Confidence")
        for row_idx in range(len(trace_df)):
            conf = trace_df.iloc[row_idx]["Confidence"]
            if conf == "HIGH_CONFIDENCE":
                ws_trace.write(row_idx + 1, conf_col, conf, high_fmt)
            elif conf == "MEDIUM_CONFIDENCE":
                ws_trace.write(row_idx + 1, conf_col, conf, med_fmt)
            elif conf == "LOW_CONFIDENCE":
                ws_trace.write(row_idx + 1, conf_col, conf, low_fmt)
            elif conf == "NO_MATCH":
                ws_trace.write(row_idx + 1, conf_col, conf, no_fmt)

        # Column widths
        ws_trace.set_column(0, 0, 14)   # Portal ID
        ws_trace.set_column(1, 1, 20)   # Portal Category
        ws_trace.set_column(2, 2, 40)   # Portal Title
        ws_trace.set_column(3, 3, 50)   # Portal Description
        ws_trace.set_column(4, 4, 30)   # 3P Sheet
        ws_trace.set_column(5, 5, 25)   # 3P Category
        ws_trace.set_column(6, 6, 14)   # 3P ID
        ws_trace.set_column(7, 7, 40)   # 3P Title
        ws_trace.set_column(8, 8, 40)   # 3P Preconditions
        ws_trace.set_column(9, 9, 50)   # 3P Steps
        ws_trace.set_column(10, 10, 50) # 3P Steps Separated
        ws_trace.set_column(11, 11, 20) # Confidence
        ws_trace.set_column(12, 12, 14) # Score

        # ---- EXECUTIVE SUMMARY SHEET ----
        ws_summary = workbook.add_worksheet("Executive Summary")
        writer.sheets["Executive Summary"] = ws_summary

        # Formats
        title_fmt = workbook.add_format({'bold': True, 'font_size': 16, 'bottom': 2})
        section_fmt = workbook.add_format({'bold': True, 'font_size': 12, 'bg_color': '#4472C4', 'font_color': 'white'})
        label_fmt = workbook.add_format({'bold': True, 'valign': 'top'})
        value_fmt = workbook.add_format({'valign': 'top'})
        pct_fmt = workbook.add_format({'num_format': '0.0%', 'bold': True, 'font_size': 14})

        row = 0
        ws_summary.write(row, 0, "Test Traceability - Executive Summary", title_fmt)
        row += 2

        # Overall Stats
        ws_summary.write(row, 0, "OVERALL COVERAGE", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        row += 1

        ws_summary.write(row, 0, "Total PORTAL (Automation) Test Cases:", label_fmt)
        ws_summary.write(row, 1, summary["total_portal"], value_fmt)
        row += 1
        ws_summary.write(row, 0, "Total 3P (Manual) Test Cases Available:", label_fmt)
        ws_summary.write(row, 1, summary["total_3p"], value_fmt)
        row += 1
        ws_summary.write(row, 0, "Total 3P Test Cases Mapped:", label_fmt)
        ws_summary.write(row, 1, summary["total_3p_mapped"], value_fmt)
        row += 1
        ws_summary.write(row, 0, "3P Coverage Percentage:", label_fmt)
        ws_summary.write(row, 1, summary["coverage_pct"] / 100, pct_fmt)
        row += 2

        # Coverage Breakdown
        ws_summary.write(row, 0, "PORTAL TEST CASE COVERAGE BREAKDOWN", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        row += 1

        ws_summary.write(row, 0, "Full Coverage (HIGH confidence match):", label_fmt)
        ws_summary.write(row, 1, summary["full_coverage"], value_fmt)
        row += 1
        ws_summary.write(row, 0, "Partial Coverage (MEDIUM/LOW matches only):", label_fmt)
        ws_summary.write(row, 1, summary["partial_coverage"], value_fmt)
        row += 1
        ws_summary.write(row, 0, "No Coverage (NO_MATCH):", label_fmt)
        ws_summary.write(row, 1, summary["no_coverage"], value_fmt)
        row += 2

        # Category breakdown - PORTAL
        ws_summary.write(row, 0, "PORTAL CATEGORY COVERAGE", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        ws_summary.write(row, 3, "", section_fmt)
        row += 1

        ws_summary.write(row, 0, "Category", label_fmt)
        ws_summary.write(row, 1, "Total TCs", label_fmt)
        ws_summary.write(row, 2, "Covered", label_fmt)
        ws_summary.write(row, 3, "Coverage %", label_fmt)
        row += 1

        for cat, info in sorted(summary["category_portal"].items(), key=lambda x: x[1]["covered"], reverse=True):
            ws_summary.write(row, 0, cat, value_fmt)
            ws_summary.write(row, 1, info["total"], value_fmt)
            ws_summary.write(row, 2, info["covered"], value_fmt)
            pct = (info["covered"] / info["total"] * 100) if info["total"] > 0 else 0
            ws_summary.write(row, 3, f"{pct:.0f}%", value_fmt)
            row += 1

        row += 1

        # 3P Suite Coverage
        ws_summary.write(row, 0, "3P SUITE MAPPING DISTRIBUTION", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        row += 1

        ws_summary.write(row, 0, "3P Suite/Sheet", label_fmt)
        ws_summary.write(row, 1, "Mappings Found", label_fmt)
        row += 1

        for sheet, count in sorted(summary["three_p_cat_mapped"].items(), key=lambda x: x[1], reverse=True):
            ws_summary.write(row, 0, sheet, value_fmt)
            ws_summary.write(row, 1, count, value_fmt)
            row += 1

        row += 1

        # Gaps
        ws_summary.write(row, 0, "COVERAGE GAPS", section_fmt)
        ws_summary.write(row, 1, "", section_fmt)
        ws_summary.write(row, 2, "", section_fmt)
        row += 1

        gap_cats = [cat for cat, info in summary["category_portal"].items()
                    if info["covered"] < info["total"]]
        if gap_cats:
            ws_summary.write(row, 0, "Categories with incomplete coverage:", label_fmt)
            row += 1
            for cat in gap_cats:
                info = summary["category_portal"][cat]
                gap = info["total"] - info["covered"]
                ws_summary.write(row, 0, f"  • {cat}: {gap} of {info['total']} TCs not fully covered", value_fmt)
                row += 1
        else:
            ws_summary.write(row, 0, "All categories have coverage.", value_fmt)

        # Set column widths for summary
        ws_summary.set_column(0, 0, 45)
        ws_summary.set_column(1, 1, 18)
        ws_summary.set_column(2, 2, 15)
        ws_summary.set_column(3, 3, 15)

    print(f"Output written to: {OUTPUT_FILE}")


def main():
    print("=" * 60)
    print("TRACEABILITY MAPPER")
    print("=" * 60)

    # Load test cases
    print("\n--- Loading PORTAL test cases ---")
    portal_tcs = load_portal_test_cases()

    print("\n--- Loading 3P test cases ---")
    three_p_tcs = load_3p_test_cases()

    # Compute mappings
    print("\n--- Computing mappings ---")
    mappings = compute_mappings(portal_tcs, three_p_tcs)

    # Generate summary
    print("\n--- Generating executive summary ---")
    summary = generate_executive_summary(mappings, portal_tcs, three_p_tcs)

    # Write output
    write_output(mappings, summary, portal_tcs, three_p_tcs)

    # Print quick summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"PORTAL test cases: {summary['total_portal']}")
    print(f"3P test cases mapped: {summary['total_3p_mapped']} / {summary['total_3p']}")
    print(f"Coverage: {summary['coverage_pct']:.1f}%")
    print(f"Full coverage: {summary['full_coverage']}")
    print(f"Partial coverage: {summary['partial_coverage']}")
    print(f"No coverage: {summary['no_coverage']}")
    print(f"\nOutput file: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
